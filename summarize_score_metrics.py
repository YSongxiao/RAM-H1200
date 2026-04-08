import csv
from pathlib import Path

from openpyxl import Workbook


EXP_PATHS = [
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_convnextv2_20260406214233",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_densenet_20260406043912",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_efficientformer_20260406124424",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_efficientnetv2_20260407010854",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_levit_20260406151143",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_mambavisiont_20260406192728",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_medmamba_20260406065848",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_mobilevit_20260406171657",
    "/mnt/data1/songxiao/RAM-H1200/ckpts/Baseline_BEScore_be_resnet34_20260406034451",
]
OUTPUT_PATH = "/mnt/data1/songxiao/RAM-H1200/ckpts/be_score_test_metrics_summary.xlsx"
DECIMALS = 4


def infer_model_name(exp_dir: Path) -> str:
    name = exp_dir.name
    if name.startswith("Baseline_BEScore_be_"):
        suffix = name[len("Baseline_BEScore_be_"):]
        return suffix.rsplit("_", 1)[0]
    return name


def resolve_csv_path(exp_path_str: str) -> Path:
    exp_path = Path(exp_path_str)
    if exp_path.is_file():
        return exp_path
    csv_path = exp_path / "test_metrics.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"test_metrics.csv not found under {exp_path}")
    return csv_path


def read_score_metrics(csv_path: Path):
    with csv_path.open("r", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    if not rows:
        raise ValueError(f"No rows found in {csv_path}")
    if "Scope" not in fieldnames or "Name" not in fieldnames:
        raise ValueError(f"Expected 'Scope' and 'Name' columns in {csv_path}")

    overall_row = None
    joint_rows = []
    metric_columns = [col for col in fieldnames if col not in ("Scope", "Name")]

    for row in rows:
        scope = str(row["Scope"]).strip()
        if scope == "Overall":
            overall_row = row
        elif scope == "Joint":
            joint_rows.append(row)

    if overall_row is None:
        raise ValueError(f"No Overall row found in {csv_path}")

    return overall_row, joint_rows, metric_columns


def to_number(value):
    text = str(value).strip()
    if text == "":
        return ""
    try:
        return float(text)
    except ValueError:
        return text


def build_overall_rows(experiments):
    metric_columns = []
    for exp in experiments:
        for col in exp["metric_columns"]:
            if col not in metric_columns:
                metric_columns.append(col)

    rows = [["Model", "Experiment", "CSV Path"] + metric_columns]
    for exp in experiments:
        overall_row = exp["overall_row"]
        row = [exp["model"], exp["experiment"], str(exp["csv_path"])]
        row.extend(to_number(overall_row.get(col, "")) for col in metric_columns)
        rows.append(row)
    return rows


def build_joint_rows(experiments):
    joint_names = []
    metric_columns = []
    for exp in experiments:
        for joint_row in exp["joint_rows"]:
            joint_name = joint_row["Name"]
            if joint_name not in joint_names:
                joint_names.append(joint_name)
        for col in exp["metric_columns"]:
            if col not in metric_columns:
                metric_columns.append(col)

    header = ["Model", "Experiment", "CSV Path"]
    for joint_name in joint_names:
        for metric_name in metric_columns:
            header.append(f"{joint_name} {metric_name}")

    rows = [header]
    for exp in experiments:
        row = [exp["model"], exp["experiment"], str(exp["csv_path"])]
        joint_map = {joint_row["Name"]: joint_row for joint_row in exp["joint_rows"]}
        for joint_name in joint_names:
            joint_row = joint_map.get(joint_name, {})
            for metric_name in metric_columns:
                row.append(to_number(joint_row.get(metric_name, "")))
        rows.append(row)
    return rows


def autosize_worksheet(ws):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 40)


def write_excel(output_path: Path, overall_rows, joint_rows, decimals: int):
    wb = Workbook()
    ws_overall = wb.active
    ws_overall.title = "Overall"
    for row in overall_rows:
        ws_overall.append(row)

    ws_joint = wb.create_sheet("Joint")
    for row in joint_rows:
        ws_joint.append(row)

    number_format = "0." + ("0" * decimals)
    for ws in (ws_overall, ws_joint):
        for row in ws.iter_rows(min_row=2):
            for cell in row[3:]:
                if isinstance(cell.value, float):
                    cell.number_format = number_format
        autosize_worksheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def print_console_summary(experiments, decimals: int):
    for exp in experiments:
        overall = exp["overall_row"]
        print(f"\n[{exp['experiment']}]")
        print(f"Model: {exp['model']}")
        compact = []
        for metric_name in ("Accuracy", "F1score", "QWK", "MAE", "Within1", "Pos/Neg ACC"):
            if metric_name in exp["metric_columns"]:
                value = to_number(overall.get(metric_name, ""))
                if isinstance(value, float):
                    compact.append(f"{metric_name}: {value:.{decimals}f}")
        print(" | ".join(compact))


def main():
    experiments = []

    for exp_path_str in EXP_PATHS:
        csv_path = resolve_csv_path(exp_path_str)
        overall_row, joint_rows, metric_columns = read_score_metrics(csv_path)
        exp_dir = csv_path.parent
        experiments.append(
            {
                "model": infer_model_name(exp_dir),
                "experiment": exp_dir.name,
                "csv_path": csv_path,
                "overall_row": overall_row,
                "joint_rows": joint_rows,
                "metric_columns": metric_columns,
            }
        )

    overall_rows = build_overall_rows(experiments)
    joint_rows = build_joint_rows(experiments)
    write_excel(Path(OUTPUT_PATH), overall_rows, joint_rows, DECIMALS)
    print_console_summary(experiments, DECIMALS)
    print(f"\nSaved Excel to: {Path(OUTPUT_PATH).resolve()}")


if __name__ == "__main__":
    main()
