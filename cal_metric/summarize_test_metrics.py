import csv
import math
from pathlib import Path

from openpyxl import Workbook


EXP_PATHS = [
    # Each entry is a segmentation experiment directory containing a `test_metrics.csv`
    # produced by `main_seg.py` / `main_be_seg.py --mode test` (bone or BE segmentation),
    # or a direct path to such a CSV file. Replace with your own runs.
    "./ckpts/Baseline_BoneSeg_swinumamba_YYYYMMDDHHMMSS",
    "./ckpts/Baseline_BoneSeg_unet_YYYYMMDDHHMMSS",
]
OUTPUT_PATH = "./bone_seg_metrics_summary.xlsx"
DECIMALS = 4


def infer_model_name(exp_dir: Path) -> str:
    name = exp_dir.name
    if name.startswith("Baseline_"):
        parts = name.split("_")
        if len(parts) >= 3:
            return parts[2]
    return name


def resolve_csv_path(exp_path_str: str) -> Path:
    exp_path = Path(exp_path_str)
    if exp_path.is_file():
        return exp_path
    csv_path = exp_path / "test_metrics.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"test_metrics.csv not found under {exp_path}")
    return csv_path


def read_metrics(csv_path: Path):
    with csv_path.open("r", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    if not rows:
        raise ValueError(f"No rows found in {csv_path}")
    if "Case" not in fieldnames:
        raise ValueError(f"'Case' column not found in {csv_path}")

    has_average_row = str(rows[-1]["Case"]).strip().lower() == "average"
    case_rows = rows[:-1] if has_average_row else rows
    mean_row = rows[-1] if has_average_row else None

    mean_columns = [col for col in fieldnames if col.startswith("Mean ")]
    class_columns = [col for col in fieldnames if col != "Case" and not col.startswith("Mean ")]

    if not mean_columns:
        raise ValueError(f"No 'Mean ' columns found in {csv_path}")

    return case_rows, mean_row, mean_columns, class_columns


def to_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def sample_std(values):
    clean = [v for v in values if not math.isnan(v)]
    if len(clean) < 2:
        return math.nan
    mean_value = sum(clean) / len(clean)
    variance = sum((v - mean_value) ** 2 for v in clean) / (len(clean) - 1)
    return math.sqrt(variance)


def summarize_columns(case_rows, mean_row, columns):
    summary = {}
    for col in columns:
        case_values = [to_float(row[col]) for row in case_rows]
        if mean_row is None:
            valid = [v for v in case_values if not math.isnan(v)]
            mean_value = sum(valid) / len(valid) if valid else math.nan
        else:
            mean_value = to_float(mean_row[col])
        summary[col] = {
            "mean": mean_value,
            "sd": sample_std(case_values),
        }
    return summary


def build_sheet_rows(experiments, metric_type):
    all_columns = []
    for exp in experiments:
        columns = exp[f"{metric_type}_columns"]
        for col in columns:
            if col not in all_columns:
                all_columns.append(col)

    header = ["Model", "Experiment", "CSV Path"]
    for col in all_columns:
        header.append(f"{col} Mean")
        header.append(f"{col} SD")

    rows = [header]
    for exp in experiments:
        row = [exp["model"], exp["experiment"], str(exp["csv_path"])]
        summary = exp[f"{metric_type}_summary"]
        for col in all_columns:
            stats = summary.get(col)
            if stats is None:
                row.extend(["", ""])
            else:
                row.extend([stats["mean"], stats["sd"]])
        rows.append(row)
    return rows


def autosize_worksheet(ws):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 50)



def write_excel(output_path: Path, mean_rows, class_rows, decimals: int):
    wb = Workbook()
    ws_mean = wb.active
    ws_mean.title = "Mean"
    for row in mean_rows:
        ws_mean.append(row)

    ws_class = wb.create_sheet("PerClass")
    for row in class_rows:
        ws_class.append(row)

    number_format = "0." + ("0" * decimals)
    for ws in (ws_mean, ws_class):
        for row in ws.iter_rows(min_row=2):
            for cell in row[3:]:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    if not math.isnan(cell.value):
                        cell.number_format = number_format
        autosize_worksheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)



def print_console_summary(experiments, decimals: int):
    for exp in experiments:
        print(f"\n[{exp['experiment']}]")
        print(f"Model: {exp['model']}")
        compact = []
        for col in exp["mean_columns"]:
            stats = exp["mean_summary"][col]
            compact.append(
                f"{col.replace('Mean ', '')}: "
                f"{stats['mean']:.{decimals}f} ± {stats['sd']:.{decimals}f}"
            )
        print(" | ".join(compact))



def main():
    experiments = []

    for exp_path_str in EXP_PATHS:
        csv_path = resolve_csv_path(exp_path_str)
        exp_dir = csv_path.parent
        case_rows, mean_row, mean_columns, class_columns = read_metrics(csv_path)

        experiments.append(
            {
                "model": infer_model_name(exp_dir),
                "experiment": exp_dir.name,
                "csv_path": csv_path,
                "mean_columns": mean_columns,
                "class_columns": class_columns,
                "mean_summary": summarize_columns(case_rows, mean_row, mean_columns),
                "class_summary": summarize_columns(case_rows, mean_row, class_columns),
            }
        )

    mean_rows = build_sheet_rows(experiments, "mean")
    class_rows = build_sheet_rows(experiments, "class")
    write_excel(Path(OUTPUT_PATH), mean_rows, class_rows, DECIMALS)
    print_console_summary(experiments, DECIMALS)
    print(f"\nSaved Excel to: {Path(OUTPUT_PATH).resolve()}")


if __name__ == "__main__":
    main()
